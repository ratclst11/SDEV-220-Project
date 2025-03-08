""" 
Shaun Ratcliff
3/8/25  
SDEV 220 Final Project
Program will be used by a manufacturing company. The system will use a GUI for entering information. 
Enter data from the alignment of a product at the factory for future field support, if any issues occur when the product is put into service. 
"""


import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

class Product:
    def __init__(self, serial_number, product_eso, date_of_build, technician_id, shim_rf, shim_rr, shim_lf, shim_lr):
        self.serial_number = serial_number
        self.product_eso = product_eso
        self.date_of_build = date_of_build
        self.technician_id = technician_id
        self.shim_rf = shim_rf
        self.shim_rr = shim_rr
        self.shim_lf = shim_lf
        self.shim_lr = shim_lr

class Gap:
    def __init__(self, gap_3, gap_6, gap_9, gap_12):
        self.gaps = [gap_3, gap_6, gap_9, gap_12]

    def validate_gaps(self):
        for gap in self.gaps:
            if not (0.0000 <= gap <= 0.0762):
                raise ValueError(f"Gap thickness out of range: {gap}")

class ExcelHandler:
    def __init__(self, file_path):
        self.file_path = file_path
        self.headers = ["Technician ID", "Serial Number", "Product ESO", "Date of Build", 
                        "Gap at 3 o'clock (mm)", "Gap at 6 o'clock (mm)", "Gap at 9 o'clock (mm)", "Gap at 12 o'clock (mm)",
                        "Shim Thickness RF (mm)", "Shim Thickness RR (mm)", "Shim Thickness LF (mm)", "Shim Thickness LR (mm)"]
        print(f"File path: {self.file_path}")  # Debugging statement

    def save_data(self, product, gap):
        try:
            if os.path.exists(self.file_path):
                workbook = load_workbook(self.file_path)
                sheet = workbook.active
                if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
                    sheet.append(self.headers)
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(self.headers)
            
            # Debugging: Print the data being saved
            print(f"Saving data: {[product.technician_id, product.serial_number, product.product_eso, product.date_of_build] + gap.gaps + [product.shim_rf, product.shim_rr, product.shim_lf, product.shim_lr]}")
            
            sheet.append([product.technician_id, product.serial_number, product.product_eso, product.date_of_build] + gap.gaps +
                         [product.shim_rf, product.shim_rr, product.shim_lf, product.shim_lr])
            workbook.save(self.file_path)
            error_label.config(text="Data saved to product_info.xlsx")
        except PermissionError:
            error_label.config(text="Permission denied: Unable to save the file. Please check your permissions.")
        except Exception as e:
            error_label.config(text=f"An error occurred: {e}")

def clear_fields():
    for entry in entries.values():
        entry.delete(0, tk.END)
    error_label.config(text="")

def submit():
    values = {label: entry.get().strip() for label, entry in entries.items()}
    
    if not all(values.values()):
        error_label.config(text="Error: All fields must be filled.")
        return
    
    try:
        datetime.strptime(values["Date of Build (YYYY-MM-DD):"], "%Y-%m-%d")
    except ValueError:
        error_label.config(text="Error: Date must be in YYYY-MM-DD format.")
        return
    
    if not values["Technician ID:"].isdigit():
        error_label.config(text="Error: Technician ID must be a numeric value.")
        return
    
    try:
        shim_values = [float(values["Shim Thickness RF (mm):"]), float(values["Shim Thickness RR (mm):"]), float(values["Shim Thickness LF (mm):"]), float(values["Shim Thickness LR (mm):"]) ]
        if not all(0.00 <= shim <= 12.7 for shim in shim_values):
            error_label.config(text="Error: Shim thickness must be between 0.00mm and 12.7mm.")
            return
    except ValueError:
        error_label.config(text="Error: Shim values must be numeric.")
        return
    
    product = Product(values["Serial Number:"], values["Product ESO:"], values["Date of Build (YYYY-MM-DD):"],
                      values["Technician ID:"], *shim_values)
    
    try:
        gap = Gap(float(values["Gap at 3 o'clock (mm):"]), float(values["Gap at 6 o'clock (mm):"]),
                  float(values["Gap at 9 o'clock (mm):"]), float(values["Gap at 12 o'clock (mm):"]))
        gap.validate_gaps()
        error_label.config(text="")
    except ValueError as e:
        error_label.config(text=f"Error: {e}. Please correct the values.")
        return
    
    excel_handler = ExcelHandler(os.path.expanduser(r"C:\Users\srat\OneDrive\Documents\product entry.xlsx"))
    excel_handler.save_data(product, gap)

root = tk.Tk()
root.title("Product Information Entry")

fields = [
    ("Technician ID:", 0),
    ("Serial Number:", 1),
    ("Product ESO:", 2),
    ("Date of Build (YYYY-MM-DD):", 3),
    ("Gap at 3 o'clock (mm):", 4),
    ("Gap at 6 o'clock (mm):", 5),
    ("Gap at 9 o'clock (mm):", 6),
    ("Gap at 12 o'clock (mm):", 7),
    ("Shim Thickness RF (mm):", 8),
    ("Shim Thickness RR (mm):", 9),
    ("Shim Thickness LF (mm):", 10),
    ("Shim Thickness LR (mm):", 11),
]

entries = {}
for label, row in fields:
    ttk.Label(root, text=label).grid(column=0, row=row, padx=10, pady=5)
    entry = ttk.Entry(root)
    entry.grid(column=1, row=row, padx=10, pady=5)
    entries[label] = entry

submit_button = ttk.Button(root, text="Submit", command=submit)
submit_button.grid(column=0, row=12, columnspan=2, pady=10)

clear_button = ttk.Button(root, text="Clear", command=clear_fields)
clear_button.grid(column=0, row=13, columnspan=2, pady=10)

error_label = ttk.Label(root, text="", foreground="red")
error_label.grid(column=0, row=14, columnspan=2, pady=5)

root.mainloop()