"""
Shaun Ratcliff
2/22/25
SDEV 220 Final Project
Alignment Checks
"""
import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

class Product:
    def __init__(self, serial_number, product_eso, date_of_build):
        self.serial_number = serial_number
        self.product_eso = product_eso
        self.date_of_build = date_of_build

class Gap:
    def __init__(self, gap_3, gap_6, gap_9, gap_12):
        self.gaps = [gap_3, gap_6, gap_9, gap_12]

    def validate_gaps(self):
        for gap in self.gaps:
            if not (0.0000 <= gap <= 0.0762):
                raise ValueError(f"Gap thickness out of range: {gap}")

class ExcelHandler:
    def __init__(self, file_path):
        self.file_path = file_path

    def save_data(self, product, gap):
        try:
            if os.path.exists(self.file_path):
                workbook = load_workbook(self.file_path)
                sheet = workbook.active
                # Check if headers are present
                if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(row=1, column=1).value is None:
                    sheet.append(["Serial Number", "Product ESO", "Date of Build", "Gap at 3 o'clock (mm)", "Gap at 6 o'clock (mm)", "Gap at 9 o'clock (mm)", "Gap at 12 o'clock (mm)"])
            else:
                workbook = Workbook()
                sheet = workbook.active
                # Create headers if the file is new
                sheet.append(["Serial Number", "Product ESO", "Date of Build", "Gap at 3 o'clock (mm)", "Gap at 6 o'clock (mm)", "Gap at 9 o'clock (mm)", "Gap at 12 o'clock (mm)"])
            
            # Append the data to the spreadsheet
            sheet.append([product.serial_number, product.product_eso, product.date_of_build] + gap.gaps)
            workbook.save(self.file_path)
            
            error_label.config(text="Data saved to product_info.xlsx")
        except PermissionError:
            error_label.config(text="Permission denied: Unable to save the file. Please check your permissions.")

def reset_invalid_gap(entry_widget):
    entry_widget.delete(0, tk.END)

def submit():
    serial_number = serial_number_entry.get()
    product_eso = product_eso_entry.get()
    date_of_build = date_of_build_entry.get()
    gap_3 = gap_3_entry.get()
    gap_6 = gap_6_entry.get()
    gap_9 = gap_9_entry.get()
    gap_12 = gap_12_entry.get()
    
    # Validate all fields are filled
    if not all([serial_number, product_eso, date_of_build, gap_3, gap_6, gap_9, gap_12]):
        error_label.config(text="Error: All fields must be filled.")
        return
    
    # Validate date format
    try:
        datetime.strptime(date_of_build, "%Y-%m-%d")
    except ValueError:
        error_label.config(text="Error: Date must be in YYYY-MM-DD format.")
        return
    
    # Create Product and Gap objects
    product = Product(serial_number, product_eso, date_of_build)
    gap = Gap(float(gap_3), float(gap_6), float(gap_9), float(gap_12))
    
    # Validate gap thickness
    try:
        gap.validate_gaps()
        error_label.config(text="")  # Clear error message if validation passes
    except ValueError as e:
        error_label.config(text=f"Error: {e}. Please correct the values.")
        for gap_value, entry in zip(gap.gaps, [gap_3_entry, gap_6_entry, gap_9_entry, gap_12_entry]):
            if not (0.0000 <= gap_value <= 0.0762):
                reset_invalid_gap(entry)
        return
    
    # Save data to Excel
    excel_handler = ExcelHandler(os.path.expanduser(r"C:\Users\srat\OneDrive\Documents\product info entry.xlsx"))
    excel_handler.save_data(product, gap)

# Create the main window
root = tk.Tk()
root.title("Product Information Entry")

# Create and place the labels and entry widgets
ttk.Label(root, text="Serial Number:").grid(column=0, row=0, padx=10, pady=5)
serial_number_entry = ttk.Entry(root)
serial_number_entry.grid(column=1, row=0, padx=10, pady=5)

ttk.Label(root, text="Product ESO:").grid(column=0, row=1, padx=10, pady=5)
product_eso_entry = ttk.Entry(root)
product_eso_entry.grid(column=1, row=1, padx=10, pady=5)

ttk.Label(root, text="Date of Build (YYYY-MM-DD):").grid(column=0, row=2, padx=10, pady=5)
date_of_build_entry = ttk.Entry(root)
date_of_build_entry.grid(column=1, row=2, padx=10, pady=5)

# New labels and entry widgets for gap thickness
ttk.Label(root, text="Gap at 3 o'clock (mm):").grid(column=0, row=3, padx=10, pady=5)
gap_3_entry = ttk.Entry(root)
gap_3_entry.grid(column=1, row=3, padx=10, pady=5)

ttk.Label(root, text="Gap at 6 o'clock (mm):").grid(column=0, row=4, padx=10, pady=5)
gap_6_entry = ttk.Entry(root)
gap_6_entry.grid(column=1, row=4, padx=10, pady=5)

ttk.Label(root, text="Gap at 9 o'clock (mm):").grid(column=0, row=5, padx=10, pady=5)
gap_9_entry = ttk.Entry(root)
gap_9_entry.grid(column=1, row=5, padx=10, pady=5)

ttk.Label(root, text="Gap at 12 o'clock (mm):").grid(column=0, row=6, padx=10, pady=5)
gap_12_entry = ttk.Entry(root)
gap_12_entry.grid(column=1, row=6, padx=10, pady=5)

# Create and place the submit button at the bottom
submit_button = ttk.Button(root, text="Submit", command=submit)
submit_button.grid(column=0, row=7, columnspan=2, pady=10)

# Create and place the error label
error_label = ttk.Label(root, text="", foreground="red")
error_label.grid(column=0, row=8, columnspan=2, pady=5)

# Run the application
root.mainloop()